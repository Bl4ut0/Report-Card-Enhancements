import openpyxl
import os
import sys

def compare_sheets(sheet1, sheet2):
    diffs = []
    max_r = max(sheet1.max_row, sheet2.max_row)
    max_c = max(sheet1.max_column, sheet2.max_column)
    
    for r in range(1, max_r + 1):
        for c in range(1, max_c + 1):
            val1 = sheet1.cell(row=r, column=c).value
            val2 = sheet2.cell(row=r, column=c).value
            
            # Normalize numeric types to avoid float comparison issues
            if isinstance(val1, float) and isinstance(val2, float):
                if abs(val1 - val2) < 0.01:
                    continue
            if val1 != val2:
                # Format cell coordinate (e.g. A1, B5)
                coord = f"{openpyxl.utils.get_column_letter(c)}{r}"
                diffs.append((coord, val1, val2))
    return diffs

def compare_workbooks(file1, file2, label):
    print(f"\n==========================================")
    print(f"Comparing {label} Workbooks:")
    print(f"  File A (Original): {os.path.basename(file1)}")
    print(f"  File B (Complete): {os.path.basename(file2)}")
    print(f"==========================================")
    
    if not os.path.exists(file1):
        print(f"Error: {file1} not found")
        return
    if not os.path.exists(file2):
        print(f"Error: {file2} not found")
        return
        
    wb1 = openpyxl.load_workbook(file1, data_only=True)
    wb2 = openpyxl.load_workbook(file2, data_only=True)
    
    sheets1 = set(wb1.sheetnames)
    sheets2 = set(wb2.sheetnames)
    
    common_sheets = sheets1.intersection(sheets2)
    only_in_1 = sheets1 - sheets2
    only_in_2 = sheets2 - sheets1
    
    if only_in_1:
        print(f"[WARN] Sheets only in Original: {only_in_1}")
    if only_in_2:
        print(f"[WARN] Sheets only in Complete: {only_in_2}")
        
    for name in sorted(common_sheets):
        sheet1 = wb1[name]
        sheet2 = wb2[name]
        
        diffs = compare_sheets(sheet1, sheet2)
        if not diffs:
            print(f"[PASS] Sheet '{name}': Perfect Match")
        else:
            print(f"[FAIL] Sheet '{name}': {len(diffs)} differences found")
            # Print first 10 differences
            for coord, v1, v2 in diffs[:10]:
                print(f"   Cell {coord}: Original = '{v1}', Complete = '{v2}'")
            if len(diffs) > 10:
                print(f"   ... and {len(diffs) - 10} more differences")

if __name__ == "__main__":
    # Dynamically find repo root relative to the script location
    root_dir = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    
    # Check if arguments are passed
    if len(sys.argv) > 1:
        if len(sys.argv) == 3:
            file1 = sys.argv[1]
            file2 = sys.argv[2]
            label = "Comparison"
            if "rpb" in file1.lower() or "rpb" in file2.lower():
                label = "RPB"
            elif "cla" in file1.lower() or "cla" in file2.lower():
                label = "CLA"
            compare_workbooks(file1, file2, label)
        else:
            print("Usage:")
            print("  python tests/compare_excel.py <file_original.xlsx> <file_complete.xlsx>")
            print("Or run without arguments to automatically match files in the project root.")
            sys.exit(1)
    else:
        # Scan root directory for excel files to compare
        files = [f for f in os.listdir(root_dir) if f.endswith(".xlsx")]
        
        # Helper to find matching pairs
        def find_pair(prefix_orig, prefix_comp):
            orig = None
            comp = None
            for f in files:
                if f.startswith(prefix_comp):
                    comp = os.path.join(root_dir, f)
                elif f.startswith(prefix_orig):
                    orig = os.path.join(root_dir, f)
            return orig, comp

        cla_orig, cla_comp = find_pair("CLA for", "Complete - CLA")
        rpb_orig, rpb_comp = find_pair("RPB for", "Complete RPB")
        
        compared = False
        
        if cla_orig and cla_comp:
            compare_workbooks(cla_orig, cla_comp, "CLA")
            compared = True
        else:
            # Check for generic/fallback CLA name matches if not specific
            cla_orig_alt = next((os.path.join(root_dir, f) for f in files if "cla" in f.lower() and "complete" not in f.lower()), None)
            cla_comp_alt = next((os.path.join(root_dir, f) for f in files if "cla" in f.lower() and "complete" in f.lower()), None)
            if cla_orig_alt and cla_comp_alt:
                compare_workbooks(cla_orig_alt, cla_comp_alt, "CLA")
                compared = True

        if rpb_orig and rpb_comp:
            compare_workbooks(rpb_orig, rpb_comp, "RPB")
            compared = True
        else:
            rpb_orig_alt = next((os.path.join(root_dir, f) for f in files if "rpb" in f.lower() and "complete" not in f.lower()), None)
            rpb_comp_alt = next((os.path.join(root_dir, f) for f in files if "rpb" in f.lower() and "complete" in f.lower()), None)
            if rpb_orig_alt and rpb_comp_alt:
                compare_workbooks(rpb_orig_alt, rpb_comp_alt, "RPB")
                compared = True
                
        if not compared:
            print("No Excel workbook pairs found in the root directory for auto-matching.")
            print("\nTo compare your own sheets:")
            print("  1. Export your sheet with V1 credentials (e.g., 'CLA_V1.xlsx')")
            print("  2. Export the same sheet with V2 credentials (e.g., 'CLA_V2.xlsx')")
            print("  3. Place them in the root directory and run:")
            print("     python tests/compare_excel.py CLA_V1.xlsx CLA_V2.xlsx")
